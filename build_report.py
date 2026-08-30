#!/usr/bin/env python3
"""
build_report.py — Weekly cross-system product report

Ingests three per-system exports (Shopify orders, warehouse stock,
returns) that each identify products with a differently-formatted SKU,
reconciles them into one clean per-product summary, and writes a
formatted Excel workbook with a "Flags" sheet documenting every
judgment call made along the way.

USAGE:
    python build_report.py <orders.csv> <warehouse.csv> <returns.csv> [-o output.xlsx]

DESIGN NOTES (read this before changing column mappings):

  - Each system's CSV columns are matched by normalized name (lowercased,
    stripped of spaces/underscores), not exact string, so a header like
    "Order_ID" vs "order id" still matches. If a system renames a column
    to something unrecognizable, this script will WARN and continue with
    nulls rather than crashing — check the console output each run.

  - The three systems format the shared product identifier differently
    ("SMR-1001", "1001", "SMR_1001"). The join key is the trailing digit
    run extracted from whatever string is given (see `extract_sku_key`).
    If a future system introduces two products that differ only in a
    prefix but share the same digits, this will incorrectly merge them —
    that's a real risk worth knowing about, not just a hypothetical.

  - Column name -> canonical field mappings live in the CONFIG section
    below. Update these first if a system changes its export format.
"""

import argparse
import re
import sys
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ============================== CONFIG ==============================
# Map of canonical field name -> list of acceptable source column names
# (normalized: lowercased, spaces/underscores/hyphens stripped).
# Add alternate names here if a system's export changes.

ORDERS_COLUMNS = {
    "order_id": ["orderid", "order"],
    "sku_raw":  ["sku", "productsku", "itemsku"],
    "product":  ["product", "productname", "itemname", "description"],
    "qty":      ["qty", "quantity", "units", "unitssold"],
    "revenue":  ["revenue", "sales", "amount", "total"],
}

WAREHOUSE_COLUMNS = {
    "sku_raw":   ["itemcode", "sku", "productsku"],
    "on_hand":   ["onhand", "qtyonhand", "stock", "quantity"],
    "location":  ["location", "warehouse", "bin"],
}

RETURNS_COLUMNS = {
    "rma":     ["rma", "rmaid", "returnid"],
    "sku_raw": ["sku", "productsku", "itemsku"],
    "reason":  ["reason", "returnreason"],
    "units":   ["units", "qty", "quantity"],
}

SKU_DISPLAY_PREFIX = "SMR"  # canonical prefix used when rebuilding a display SKU


# ========================== NORMALIZATION ============================

def normalize_colname(name: str) -> str:
    """Lowercase and strip spaces/underscores/hyphens for fuzzy column matching."""
    return re.sub(r"[\s_\-]+", "", str(name).strip().lower())


def map_columns(df: pd.DataFrame, mapping: dict, source_name: str) -> pd.DataFrame:
    """
    Rename df's columns to canonical names based on `mapping`, matching
    fuzzily (case/space/underscore-insensitive). Missing expected columns
    are created as all-null and a warning is printed, rather than raising.
    """
    normalized_lookup = {normalize_colname(c): c for c in df.columns}
    rename = {}
    missing = []

    for canonical, aliases in mapping.items():
        found = None
        for alias in [canonical] + aliases:
            key = normalize_colname(alias)
            if key in normalized_lookup:
                found = normalized_lookup[key]
                break
        if found:
            rename[found] = canonical
        else:
            missing.append(canonical)

    out = df.rename(columns=rename)
    for col in missing:
        out[col] = pd.NA
    if missing:
        print(f"  [WARN] {source_name}: could not find column(s) {missing} — "
              f"filled with blanks. Check if the export format changed.")

    keep = list(mapping.keys())
    return out[keep]


def extract_sku_key(raw) -> str | None:
    """
    Pull the trailing run of digits out of a raw SKU/item-code string and
    zero-pad it to 4 digits. This is the join key across all three systems,
    since 'SMR-1001', '1001', and 'SMR_1001' all reduce to '1001'.
    Returns None if no digits are found (e.g. blank SKU).
    """
    if pd.isna(raw):
        return None
    match = re.search(r"(\d+)\s*$", str(raw).strip())
    if not match:
        return None
    return match.group(1).zfill(4)


def display_sku(key: str) -> str:
    """Rebuild a canonical display SKU from the join key, e.g. '1001' -> 'SMR-1001'."""
    return f"{SKU_DISPLAY_PREFIX}-{key}"


def to_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


# ============================= LOADERS ===============================

def load_orders(path: str, flags: list) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str)
    df = map_columns(df, ORDERS_COLUMNS, "Orders (system1)")

    df["sku_key"] = df["sku_raw"].apply(extract_sku_key)
    df["qty"] = to_numeric(df["qty"])
    df["revenue"] = to_numeric(df["revenue"])

    # Flag rows with no resolvable SKU — can't be joined to a product.
    unmapped = df[df["sku_key"].isna()]
    for _, row in unmapped.iterrows():
        flags.append({
            "sheet": "Orders",
            "issue": "Missing/unreadable SKU — order excluded from product summary",
            "detail": f"order_id={row.get('order_id')}, product={row.get('product')}",
        })

    # Flag rows with missing qty (treated as 0 in totals so revenue isn't lost).
    missing_qty = df[df["qty"].isna() & df["sku_key"].notna()]
    for _, row in missing_qty.iterrows():
        flags.append({
            "sheet": "Orders",
            "issue": "Missing quantity — treated as 0 in unit totals",
            "detail": f"order_id={row.get('order_id')}, sku={row.get('sku_raw')}",
        })
    df["qty"] = df["qty"].fillna(0)

    return df


def load_warehouse(path: str, flags: list) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str)
    df = map_columns(df, WAREHOUSE_COLUMNS, "Warehouse (system2)")

    # Exact-duplicate rows (same sku_raw + on_hand + location) are dropped.
    before = len(df)
    dupe_mask = df.duplicated(keep="first")
    for _, row in df[dupe_mask].iterrows():
        flags.append({
            "sheet": "Warehouse",
            "issue": "Exact duplicate row — dropped",
            "detail": f"item_code={row.get('sku_raw')}, on_hand={row.get('on_hand')}",
        })
    df = df[~dupe_mask].copy()

    df["sku_key"] = df["sku_raw"].apply(extract_sku_key)
    df["on_hand"] = to_numeric(df["on_hand"])

    missing_stock = df[df["on_hand"].isna() & df["sku_key"].notna()]
    for _, row in missing_stock.iterrows():
        flags.append({
            "sheet": "Warehouse",
            "issue": "Missing on-hand quantity — shown as 'Unknown', NOT assumed zero",
            "detail": f"item_code={row.get('sku_raw')}, location={row.get('location')}",
        })

    return df


def load_returns(path: str, flags: list) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str)
    df = map_columns(df, RETURNS_COLUMNS, "Returns (system3)")

    df["sku_key"] = df["sku_raw"].apply(extract_sku_key)
    df["units"] = to_numeric(df["units"]).fillna(0)

    missing = df[df["sku_key"].isna()]
    for _, row in missing.iterrows():
        flags.append({
            "sheet": "Returns",
            "issue": "Missing/unreadable SKU — return excluded from product summary",
            "detail": f"rma={row.get('rma')}",
        })

    return df


# ============================ RECONCILE ===============================

def build_summary(orders: pd.DataFrame, warehouse: pd.DataFrame,
                   returns: pd.DataFrame, flags: list) -> pd.DataFrame:

    orders_ok = orders[orders["sku_key"].notna()]
    warehouse_ok = warehouse[warehouse["sku_key"].notna()]
    returns_ok = returns[returns["sku_key"].notna()]

    orders_agg = orders_ok.groupby("sku_key").agg(
        product=("product", "first"),
        units_sold=("qty", "sum"),
        revenue=("revenue", "sum"),
    )

    warehouse_agg = warehouse_ok.groupby("sku_key").agg(
        # min_count=1 so a group where every value is NaN sums to NaN, not 0 —
        # otherwise a genuinely-unknown stock level is indistinguishable from
        # a confirmed zero (pandas' default sum() of all-NaN is 0.0).
        on_hand=("on_hand", lambda s: s.sum(min_count=1)),
        location=("location", lambda s: ", ".join(sorted(set(s.dropna())))),
        _on_hand_known=("on_hand", lambda s: s.notna().any()),
    )

    returns_agg = returns_ok.groupby("sku_key").agg(
        units_returned=("units", "sum"),
        return_reasons=("reason", lambda s: "; ".join(sorted(set(s.dropna()))) or None),
    )

    all_keys = sorted(set(orders_agg.index) | set(warehouse_agg.index) | set(returns_agg.index))
    summary = pd.DataFrame(index=all_keys)
    summary = summary.join(orders_agg).join(warehouse_agg).join(returns_agg)

    summary["sku"] = [display_sku(k) for k in summary.index]

    # Product name: fall back across sources; flag if we never learn one.
    known_warehouse_keys = set(warehouse_agg.index)
    known_order_keys = set(orders_agg.index)
    known_return_keys = set(returns_agg.index)

    for key in all_keys:
        if key not in known_order_keys and (key in known_warehouse_keys or key in known_return_keys):
            flags.append({
                "sheet": "Summary",
                "issue": "SKU has no matching order record — appears only in warehouse/returns data",
                "detail": f"sku={display_sku(key)}",
            })
        if key not in known_warehouse_keys and key in known_order_keys:
            flags.append({
                "sheet": "Summary",
                "issue": "SKU sold in orders but absent from warehouse export — stock unknown",
                "detail": f"sku={display_sku(key)}",
            })
        if key in known_return_keys and key not in known_order_keys and key not in known_warehouse_keys:
            flags.append({
                "sheet": "Summary",
                "issue": "Returned SKU not found in orders or warehouse at all — likely bad SKU or discontinued product",
                "detail": f"sku={display_sku(key)}",
            })

    summary["product"] = summary["product"].fillna("(unknown — not in orders export)")
    summary["units_sold"] = summary["units_sold"].fillna(0)
    summary["revenue"] = summary["revenue"].fillna(0.0)
    summary["units_returned"] = summary["units_returned"].fillna(0)
    summary["return_reasons"] = summary["return_reasons"].fillna("")
    summary["location"] = summary["location"].fillna("")

    # Distinguish three cases for on-hand stock:
    #   - no warehouse row for this SKU at all            -> "No record"
    #   - warehouse row exists but the value was blank     -> "Unknown"
    #   - warehouse row exists with a real number          -> that number
    def stock_display(row):
        if key_in_warehouse := (row.name in known_warehouse_keys):
            if pd.isna(row["on_hand"]):
                return "Unknown"
            return int(row["on_hand"])
        return "No record"
    summary["on_hand"] = summary.apply(stock_display, axis=1)

    summary = summary[[
        "sku", "product", "units_sold", "revenue",
        "on_hand", "location", "units_returned", "return_reasons",
    ]].reset_index(drop=True)

    summary = summary.sort_values("sku").reset_index(drop=True)
    return summary


# ============================== OUTPUT ================================

def write_workbook(summary: pd.DataFrame, flags_df: pd.DataFrame, out_path: str):
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        flags_df.to_excel(writer, sheet_name="Flags", index=False)

    wb = load_workbook(out_path)
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    body_font = Font(name="Arial")

    for sheet_name in ("Summary", "Flags"):
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
        # Autofit-ish column widths
        for col_cells in ws.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max(length + 3, 12), 50)
        ws.freeze_panes = "A2"

    # Revenue column as currency (Summary column D)
    for row in wb["Summary"].iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '$#,##0.00'

    wb.save(out_path)


# ================================ MAIN =================================

def main():
    parser = argparse.ArgumentParser(description="Build the weekly cross-system product report.")
    parser.add_argument("orders_csv")
    parser.add_argument("warehouse_csv")
    parser.add_argument("returns_csv")
    parser.add_argument("-o", "--output", default="weekly_report.xlsx")
    args = parser.parse_args()

    flags: list[dict] = []

    print(f"Loading orders from {args.orders_csv} ...")
    orders = load_orders(args.orders_csv, flags)
    print(f"Loading warehouse export from {args.warehouse_csv} ...")
    warehouse = load_warehouse(args.warehouse_csv, flags)
    print(f"Loading returns from {args.returns_csv} ...")
    returns = load_returns(args.returns_csv, flags)

    print("Reconciling ...")
    summary = build_summary(orders, warehouse, returns, flags)

    flags_df = pd.DataFrame(flags) if flags else pd.DataFrame(columns=["sheet", "issue", "detail"])
    flags_df.insert(0, "run_date", datetime.now().strftime("%Y-%m-%d"))

    write_workbook(summary, flags_df, args.output)

    print(f"\nDone. {len(summary)} products in summary, {len(flags_df)} items flagged for review.")
    print(f"Output written to: {args.output}")


if __name__ == "__main__":
    sys.exit(main())
